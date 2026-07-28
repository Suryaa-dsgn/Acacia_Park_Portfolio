# scripts/ingest.py
# Governed-data ingestion for the Acacia property (build plan Phase 6 data work).
# Reads the two real workbooks and emits scripts/acacia-governed.json, the real
# governed data the fixture generator assembles into the contract:
#   - Acacia_Q1_2026_Quarterly_Narrative_SCRUBBED.xlsx: the product partner's
#     governed Q1 2026 report (statement, rent roll, leasing, occupancy 2026,
#     narrative, provenance).
#   - "Acacia - 2025 Q4 (2).xlsx": the client's Q4 2025 export (statement PTD+YTD,
#     occupancy 2024 and 2025, narrative, identity, GL codes).
#
# Run: python scripts/ingest.py   (requires openpyxl)
#
# The portal renders governed data. Where the Q4 2025 export lacks a governed
# rent schedule or leasing summary, this is recorded honestly (rent metrics
# carried from the Q1 governed snapshot and flagged; leasing left ungoverned).
import json
import os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRUBBED = os.path.join(ROOT, "Acacia_Q1_2026_Quarterly_Narrative_SCRUBBED.xlsx")
RAW_Q4 = os.path.join(ROOT, "Acacia - 2025 Q4 (2).xlsx")
OUT = os.path.join(ROOT, "scripts", "acacia-governed.json")

# Canonical statement keys and GL codes (from the Q4 chart of accounts).
INCOME = [
    ("rentalIncome", "Rental Income", "4300-0000"),
    ("otherResidentIncome", "Other Resident Income", "4499-0000"),
    ("otherPropertyIncome", "Other Property Income", "4998-0000"),
]
EXPENSES = [
    ("fixedExpenses", "Fixed Expenses", "5199-0000"),
    ("repairsMaintenance", "Repairs & Maintenance", "5399-0000"),
    ("utilities", "Utilities", "5499-0000"),
    ("gAndA", "General & Administrative", "6770-9999"),
    ("marketing", "Marketing", "6771-9999"),
    ("makeReadyTurnover", "Make Ready / Turnover", "6772-9999"),
    ("professionalServices", "Professional Services", "6776-9999"),
    ("payrollLabor", "Payroll / Labor", "6996-0000"),
]

IDENTITY = {
    "kind": "property",
    "tradeName": "Acacia Park Apartments",
    "legalEntity": "Universe at Acacia (DE), LLC",
    "manager": "Global Integrity Realty Corporation",
    "addressLine": "5280 N Little Mountain Drive",
    "cityStateZip": "San Bernardino, CA 92407",
    "submarket": "Inland Empire",
    "units": 304,
    "totalSqFt": 258400,
}


import re

# Mirror lib/format.ts normalizeText so committed fixtures carry no em dash,
# en dash, minus sign, or non-breaking space (the display layer normalizes too).
def norm(s):
    if not s:
        return s
    s = re.sub(r"\s*—\s*", ": ", s)  # em dash -> colon + space
    s = s.replace("–", "-")          # en dash -> hyphen
    s = s.replace("−", "-")          # minus sign -> hyphen
    s = s.replace(" ", " ")          # non-breaking space -> space
    s = re.sub(r"\s+", " ", s).strip()
    return s


def num(v):
    return round(float(v), 2)


def line(key, label, gl, ptd_cur, ptd_prior, ytd_cur, ytd_prior):
    return {
        "key": key,
        "label": label,
        "glCode": gl,
        "ptdCurrent": num(ptd_cur),
        "ptdPrior": num(ptd_prior),
        "ytdCurrent": num(ytd_cur),
        "ytdPrior": num(ytd_prior),
    }


# ---------------------------------------------------------------------------
# Q1 2026 from the scrubbed governed workbook
# ---------------------------------------------------------------------------
def read_q1():
    wb = openpyxl.load_workbook(SCRUBBED, data_only=True)
    ws = wb["Q1 2026 Narrative"]

    def cell(coord):
        return ws[coord].value

    # Operating statement: column C = Q1 2025 (prior), D = Q1 2026 (current).
    # Q1 has PTD == YTD (the workbook repeats them in F/G), so ytd = ptd.
    income_rows = [21, 22, 23]
    expense_rows = [27, 28, 29, 30, 31, 32, 33, 34]
    income = [
        line(k, l, gl, cell(f"D{r}"), cell(f"C{r}"), cell(f"D{r}"), cell(f"C{r}"))
        for (k, l, gl), r in zip(INCOME, income_rows)
    ]
    expenses = [
        line(k, l, gl, cell(f"D{r}"), cell(f"C{r}"), cell(f"D{r}"), cell(f"C{r}"))
        for (k, l, gl), r in zip(EXPENSES, expense_rows)
    ]

    rent_roll = {
        "asOf": "2026-03-31",
        "totalUnits": int(cell("C65")),
        "occupiedUnits": int(cell("C66")),
        "vacantUnits": int(cell("C67")),
        "physicalOccupancy": float(cell("C68")),
        "totalSqFt": int(cell("C69")),
        "avgUnitSqFt": num(cell("C70")),
        "avgUnitRent": num(cell("C71")),
        "avgOccupiedUnitRent": num(cell("C72")),
        "avgResidentRent": num(cell("C73")),
        "totalMarketRent": num(cell("C74")),
        "totalInPlaceRent": num(cell("C75")),
        "economicOccupancy": float(cell("C76")),
    }

    leasing = {
        "moveIns": int(cell("C79")),
        "moveOuts": int(cell("C80")),
        "netAbsorption": int(cell("C81")),
        "noticesToVacate": int(cell("C82")),
        "unitsRented": int(cell("C83")),
        "renewals": int(cell("C84")),
        "evictions": int(cell("C85")),
    }

    # Occupancy 2026 (Jan-May reported, rest not yet occurred -> null).
    occ = wb["Occupancy 2025 and 2026"]
    cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    occ2026 = []
    for c in cols:
        v = occ[f"{c}5"].value
        occ2026.append(round(float(v), 6) if isinstance(v, (int, float)) else None)

    # Narrative (authoring drafts with governed baseline). Tokens substituted.
    def sub(s):
        return norm(
            (s or "")
            .replace("[PROPERTY NAME]", IDENTITY["tradeName"])
            .replace("[MARKET]", "San Bernardino")
        )

    narrative = [
        {"key": "marketCommentary", "title": "Market Commentary", "status": "Gathering Info", "body": sub(cell("B88"))},
        {"key": "propertyOperations", "title": "Property Operations", "status": "Work in Progress", "body": sub(cell("B94"))},
        {"key": "conditionsRenovation", "title": "Property Conditions & Renovation Plans", "status": "Gathering Info", "body": sub(cell("B100"))},
    ]

    # Provenance: the governed audit lines (strip the leading bullet).
    prov_cells = ["B114", "B115", "B116", "B117", "B118", "B119"]
    provenance = []
    for pc in prov_cells:
        t = cell(pc)
        if t:
            provenance.append({"text": norm(t.lstrip("• ").strip())})

    return {
        "income": income,
        "expenses": expenses,
        "rentRoll": rent_roll,
        "leasing": leasing,
        "currentYear": 2026,
        "priorYear": 2025,
        "occCurrent": occ2026,  # 2026
        "occPrior": None,       # filled from the raw 2025 series below
        "narrative": narrative,
        "provenance": provenance,
    }


# ---------------------------------------------------------------------------
# Q4 2025 from the raw client export
# ---------------------------------------------------------------------------
def read_q4():
    wb = openpyxl.load_workbook(RAW_Q4, data_only=True)
    ws = wb["Q4 Narrative"]

    def cell(coord):
        return ws[coord].value

    # Column E = Q4 2024 (prior PTD), F = Q4 2025 (current PTD),
    # H = YTD 2024 (prior), I = YTD 2025 (current).
    income_rows = [34, 35, 36]
    expense_rows = [41, 42, 43, 44, 45, 46, 47, 48]
    income = [
        line(k, l, gl, cell(f"F{r}"), cell(f"E{r}"), cell(f"I{r}"), cell(f"H{r}"))
        for (k, l, gl), r in zip(INCOME, income_rows)
    ]
    expenses = [
        line(k, l, gl, cell(f"F{r}"), cell(f"E{r}"), cell(f"I{r}"), cell(f"H{r}"))
        for (k, l, gl), r in zip(EXPENSES, expense_rows)
    ]

    # Occupancy: 2024 fractions in row 13, 2025 fractions in row 28.
    occ = wb["Occupancy 2024 and 2025"]
    cols = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    occ2024 = [round(float(occ[f"{c}13"].value), 6) for c in cols]
    occ2025 = [round(float(occ[f"{c}28"].value), 6) for c in cols]

    # Average in-place rent per unit per month, on the Q4 Narrative sheet:
    # 2024 in row 89, 2025 in row 90, months in columns Z..AK (Jan..Dec).
    rent_cols = ["Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK"]
    rent2024 = [round(float(cell(f"{c}89")), 2) for c in rent_cols]
    rent2025 = [round(float(cell(f"{c}90")), 2) for c in rent_cols]

    # Real authored narrative, complete for Q4 2025.
    narrative = [
        {"key": "marketCommentary", "title": "Market Commentary", "status": "Completed", "body": norm(cell("D15"))},
        {"key": "propertyOperations", "title": "Property Operations", "status": "Completed", "body": norm(cell("D74"))},
        {"key": "conditionsRenovation", "title": "Property Conditions & Renovation Plans", "status": "Completed", "body": norm(cell("D88"))},
    ]

    return {
        "income": income,
        "expenses": expenses,
        "currentYear": 2025,
        "priorYear": 2024,
        "occCurrent": occ2025,  # 2025
        "occPrior": occ2024,    # 2024
        "occupancyRentHistory": {
            "years": [2024, 2025],
            "occupancy": [occ2024, occ2025],
            "avgInPlaceRent": [rent2024, rent2025],
        },
        "narrative": narrative,
        "provenance": [
            {"text": "Operating statement sourced from the Q4 2025 Yardi Income Statement (Cash basis), Acacia Park Apartments."},
            {"text": "Occupancy from the governed 12-month occupancy series (2024 and 2025)."},
            {"text": "Rent roll rent metrics carried from the Q1 2026 governed snapshot; the Q4 2025 rent schedule is not separately governed. Occupancy counts derived from the occupancy series."},
            {"text": "Leasing activity is not governed for Q4 2025."},
            {"text": "GL account codes from the Yardi chart of accounts."},
        ],
    }


def main():
    q1 = read_q1()
    q4 = read_q4()

    # Prior-year (2025) occupancy for the Q1 2026 trend comes from the raw file.
    q1["occPrior"] = q4["occCurrent"]

    # Q4 2025 rent roll: occupancy counts derived from the real Dec 2025
    # occupancy; rent-dollar metrics carried from the Q1 governed snapshot
    # (flagged in the Q4 provenance above).
    units = IDENTITY["units"]
    dec_occ = q4["occCurrent"][11]
    occupied = round(dec_occ * units)
    q1_rr = q1["rentRoll"]
    q4["rentRoll"] = {
        "asOf": "2025-12-31",
        "totalUnits": units,
        "occupiedUnits": occupied,
        "vacantUnits": units - occupied,
        "physicalOccupancy": round(occupied / units, 6),
        "totalSqFt": IDENTITY["totalSqFt"],
        "avgUnitSqFt": q1_rr["avgUnitSqFt"],
        "avgUnitRent": q1_rr["avgUnitRent"],
        "avgOccupiedUnitRent": q1_rr["avgOccupiedUnitRent"],
        "avgResidentRent": q1_rr["avgResidentRent"],
        "totalMarketRent": q1_rr["totalMarketRent"],
        "totalInPlaceRent": q1_rr["totalInPlaceRent"],
        "economicOccupancy": q1_rr["economicOccupancy"],
    }
    # Leasing is not governed for Q4 2025.
    q4["leasing"] = None

    data = {
        "identity": IDENTITY,
        "occupancyRentHistory": q4["occupancyRentHistory"],
        "quarters": {"2026-q1": q1, "2025-q4": q4},
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"wrote {OUT}")
    # quick sanity
    ti = sum(l["ptdCurrent"] for l in q1["income"])
    print(f"Q1 2026 Total Income (sum of lines) = {ti:.2f} (expect 1769661.34)")
    q4ti = sum(l["ptdCurrent"] for l in q4["income"])
    print(f"Q4 2025 Total Income PTD (sum) = {q4ti:.2f} (expect 1795173.53)")


if __name__ == "__main__":
    main()
